import os
import pandas as pd
from openpyxl import load_workbook

VIDEO_EXTENSIONS = ('.mp4', '.mkv', '.avi', '.mov', '.flv', '.wmv')
root_folder = 'moment'  # Đặt tên thư mục gốc ở đây

# Bước 1: Thu thập dữ liệu
data = []
for dirpath, _, filenames in os.walk(root_folder):
    folder_name = os.path.basename(dirpath)
    for file in sorted(filenames):
        if file.lower().endswith(VIDEO_EXTENSIONS):
            data.append({'Folder': folder_name, 'Video Name': file})

# Bước 2: Xuất Excel bằng pandas
df = pd.DataFrame(data)
output_file = 'thong_ke_merge_folder.xlsx'
df.to_excel(output_file, index=False)

# Bước 3: Mở bằng openpyxl và merge ô "Folder"
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

wb = load_workbook(output_file)
ws = wb.active

current_folder = None
start_row = 2  # Excel tính từ dòng 1, dòng 1 là header

for i in range(2, ws.max_row + 2):
    folder = ws[f"A{i}"].value
    if folder != current_folder:
        if current_folder is not None and i - start_row > 1:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=i - 1, end_column=1)
        current_folder = folder
        start_row = i

# Merge đoạn cuối cùng nếu cần
if ws.max_row >= start_row:
    ws.merge_cells(start_row=start_row, start_column=1, end_row=ws.max_row, end_column=1)

wb.save(output_file)
print(f"✅ Đã xuất và merge ô Folder tại: {output_file}")
